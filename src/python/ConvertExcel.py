import openpyxl as xl
from num2words import num2words
from Connection import *
import sys
import os

def eliminar_duplicados_excel(worksheet):
    # Crear un conjunto para almacenar los números de celular duplicados
    numeros_duplicados = set()
    
    # Crear una lista para almacenar los índices de las filas con datos erróneos
    indices_erroneos = []
    
    # Iterar sobre las filas del archivo Excel
    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        nombre = row[0]
        celular = row[1]
        
        # Verificar si el número de celular ya se encuentra en el conjunto de duplicados
        if celular in numeros_duplicados:
            # Verificar si el nombre es diferente al primer registro con el mismo número de celular
            indice_primero = min(i for i, r in enumerate(worksheet.iter_rows(min_row=2, max_row=index-1, values_only=True), start=2) if r[1] == celular)
            nombre_primero = worksheet.cell(row=indice_primero, column=1).value
            if nombre != nombre_primero:
                indices_erroneos.append(index)  # Guardar el índice de la fila con datos erróneos
        else:
            numeros_duplicados.add(celular)  # Agregar el número de celular al conjunto de duplicados
    
    # Eliminar las filas con datos erróneos en orden inverso
    for indice in reversed(indices_erroneos):
        worksheet.delete_rows(indice) 
        
def convertir_a_texto(valor):
    if valor >= 1000000:
        millones = int(valor / 1000000)
        resto = int(valor % 1000000)
        millones_texto = num2words(millones, lang="es").replace("-", " ")
        resto_texto = num2words(resto, lang="es").replace("-", " ")
        valor_texto = f"{millones_texto} millones {resto_texto}".upper()
    else:
        valor_texto = num2words(valor, lang="es").replace("-", " ").upper()
    return valor_texto

def procesar_lote(lote_columna_valor):
    for cell in lote_columna_valor:
        valor = cell.value
        if isinstance(valor, (int, float)):
            valor_texto = convertir_a_texto(valor)
            cell.value = valor_texto

try:    
    # Se recepciona achivo en excel y el numero de insert
    archivo_excel = sys.argv[1]  
    insertId =  sys.argv[2]  
    carpeta =  sys.argv[3]  
    print(carpeta)
    # Cargar el archivo de Excel
    workbook = xl.load_workbook(archivo_excel)
    # Obtener la hoja de trabajo activa
    sheet = workbook.active
    # Eliminar las columnas F y G
    sheet.delete_cols(6, 2)  # Eliminar desde la columna F (6) hasta la columna G (7)
    # Obtener el rango de celdas para procesar en lotes
    columna_valor = sheet['D']  # Columna "valor" (suponiendo que es la columna D)

    # Obtener los valores de la columna D y actualizar los valores sin puntos y redondeados
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            valor_original = cell.value  # Columna D
            if isinstance(valor_original, str):
                valor_sin_puntos = valor_original.replace('.', '')
                # Redondear el valor al siguiente múltiplo de 100
                valor_redondeado = round(int(valor_sin_puntos) + (100 - (int(valor_sin_puntos) % 100)))
                cell.value = valor_redondeado  # Columna D

    # Obtener los valor de la columna B
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            valor_original = cell.value  # Columna B
            if isinstance(valor_original, str):
                try:
                    valor_entero = int(valor_original)
                    cell.value = valor_entero  # Actualizar el valor en la columna B
                except ValueError:
                    # El valor no se puede convertir a entero
                    pass

    # Dividir el rango de celdas en lotes y procesar cada lote
    rango_lotes = 1000  # Cantidad de registros por lote
    total_registros = len(columna_valor) - 1  # Restar 1 para excluir el encabezado
    for lote_inicio in range(0, total_registros, rango_lotes):
        lote_fin = lote_inicio + rango_lotes
        lote_columna_valor = columna_valor[lote_inicio:lote_fin]
        procesar_lote(lote_columna_valor)

    eliminar_duplicados_excel(sheet)

    # Se crear nombre para el nuevo archivo
    nombre_archivo, extension = os.path.splitext(archivo_excel)
    archivo_nuevo = str(nombre_archivo + "_2" + extension)
    
    # Guardar el archivo modificado
    workbook.save(archivo_nuevo)

    namePath = os.path.basename(archivo_excel)
    nombre_sin_extension, extension2 = namePath.split(".")
    downloadFile = nombre_sin_extension + "_2." + extension2
    
    ruta_guardar = f"doc/base/{carpeta}/{downloadFile}"    

    ruta_final = archivo_nuevo.replace("\\", "/")

    connectionMySQL = conntDB()
    with connectionMySQL.cursor() as cursor:
        sql = f"UPDATE tbl_status SET STA_CARCHIVO_RESULTANTE = '{ruta_guardar}', STA_CPATH_RESULTANTE = '{ruta_final}' WHERE PKSTA_NCODIGO = {insertId}"
        cursor.execute(sql)    
    connectionMySQL.close

    # Imprimir mensaje de finalización
    print("Archivo modificado guardado correctamente.")

except Exception as e:
    print(e)